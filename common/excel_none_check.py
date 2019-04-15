#!/usr/bin/python3
# -*- coding:utf-8 -*-
# __author__ = 'liao gao xiang'

from openpyxl import load_workbook

wb = load_workbook(filename='stock_info.xlsx')
print('共有' + str(len(wb.sheetnames)) + '个sheet：{}'.format(wb.sheetnames))

# 1.循环workbook中所有sheet
for sheet in wb.sheetnames:
    ws = wb[sheet]
    # 2.循环每个sheet中所有单元格数据
    for row in ws.rows:
        for cell in row:
            if cell.value is None:
                try:
                    print(sheet + " " + cell.coordinate + '的数据为空：' + str(cell.value))
                    ws[cell.coordinate] = 0
                except AttributeError as e:
                    print(str(sheet) + 'sheet格式不对：' + str(e))
wb.save('stock_info.xlsx')
