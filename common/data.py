#!/usr/bin/python3
# -*- coding:utf-8 -*-
# __author__ = 'liao gao xiang'

from itertools import islice
from django.core.urlresolvers import reverse
from django.shortcuts import render, HttpResponseRedirect
from django.views.generic import View
from openpyxl import load_workbook
from stock.models import *


class BaseView(View):
    """excel解析基类"""

    def __init__(self, **kwargs):
        super(BaseView, self).__init__(**kwargs)
        self.stock_info = StockInfo.objects.all()
        self.holder_info = HolderInfo.objects.all()
        self.liquid_holder_info = LiquidHolderInfo.objects.all()
        self.market_day = MarketDay.objects.all()
        self.market_week = MarketWeek.objects.all()
        self.market_month = MarketMonth.objects.all()
        self.market_year = MarketYear.objects.all()
        self.margin = Margin.objects.all()
        self.morphologic_factor = MorphologicFactor.objects.all()
        self.technical_factor = TechnicalFactor.objects.all()
        self.financial_factor = FinancialFactor.objects.all()

    @staticmethod
    def excel(request, name):
        file = request.FILES.get(name, '')
        if str(file).split('.')[-1] not in ['xlsx', 'xlsm', 'xltx', 'xltm']:
            print("{} 格式不对！请上传.xlsx/xlsm/xltx/xltm格式结尾的文件".format(str(file)))
            return False
        else:
            wb = load_workbook(filename=file, read_only=True)  # # 只读模式加载workbook
            print('共有' + str(len(wb.sheetnames)) + '个sheet：{}'.format(wb.sheetnames))
            sheet = wb.sheetnames[0]  # 循环workbook中第一个sheet
            return wb[sheet].rows

    def get(self, request):
        return render(request, 'index.html', {})


class StockInfoView(BaseView):
    """股票信息"""

    def post(self, request):
        data = self.excel(request, "stock_info")
        if data:
            for row in islice(data, 1, None):
                try:  # 此处一定要与excel中字段顺序对应！！
                    code, name, is_st, industry, is_margin, is_hs300, is_sh, is_sz, is_sh50, is_500, is_sme, is_second, \
                    share_total, share_liqa, share_restrict, holder_num, holder_average_num, holder_average_change = \
                        map(lambda x: x.value, row)
                    code = code.split(".")[0]
                    stock_info = self.stock_info.filter(code=code)
                    if stock_info:
                        stock_info = stock_info.get()
                    else:
                        stock_info = StockInfo()
                        stock_info.code = code
                    stock_info.name = name
                    stock_info.is_st = is_st
                    stock_info.industry = industry
                    stock_info.is_margin = is_margin
                    stock_info.is_hs300 = is_hs300
                    stock_info.is_sh = is_sh
                    stock_info.is_sz = is_sz
                    stock_info.is_sh50 = is_sh50
                    stock_info.is_500 = is_500
                    stock_info.is_sme = is_sme
                    stock_info.is_second = is_second
                    stock_info.share_total = share_total
                    stock_info.share_liqa = share_liqa
                    stock_info.share_restrict = share_restrict
                    stock_info.holder_num = holder_num
                    stock_info.holder_average_num = holder_average_num
                    stock_info.holder_average_change = holder_average_change
                    try:
                        stock_info.save()
                    except Exception as e:
                        print("StockInfo表保存失败，code:{}，报错内容：{}".format(code, e))
                except Exception as e:
                    print("数据表更新失败，{}，报错内容：{}".format(row, e))
        return HttpResponseRedirect(reverse('index'))


class HolderInfoView(BaseView):
    """十大股东信息"""

    def post(self, request):
        data = self.excel(request, "holder_info")
        if data:
            for row in islice(data, 1, None):  # 表示从excel的第二行开始读取
                try:
                    list_ = []
                    list_.extend(map(lambda x: x.value, row))
                    code = list_[0].split(".")[0]  # 股票代码
                    holders = list_[2].split(";")  # 股东列表
                    quantities = list_[3: 3 + len(holders)]  # 持股数量列表
                    pcts = list_[13: 13 + len(holders)]  # 持股比例列表
                    for i in range(len(holders)):
                        holder_info = self.holder_info.filter(code=code).filter(name=holders[i])
                        if holder_info:
                            holder_info = holder_info.get()
                        else:
                            holder_info = HolderInfo()
                            holder_info.code = code
                            holder_info.name = holders[i]
                        holder_info.quantity = quantities[i]
                        holder_info.pct = pcts[i]
                        try:
                            holder_info.save()
                        except Exception as e:
                            print("HolderInfo表保存失败，code_name:{}_{}，报错内容：{}".format(code, quantities[i], e))
                except Exception as e:
                    print("数据表更新失败，{}，报错内容：{}".format(row, e))
        return HttpResponseRedirect(reverse('index'))


class LiquidHolderInfoView(BaseView):
    """十大股东信息"""

    def post(self, request):
        data = self.excel(request, "liquid_holder_info")
        if data:
            for row in islice(data, 1, None):  # 表示从excel的第二行开始读取
                try:
                    list_ = []
                    list_.extend(map(lambda x: x.value, row))
                    code = list_[0].split(".")[0]  # 股票代码
                    holders = list_[2].split(";")  # 股东列表
                    quantities = list_[3: 3 + len(holders)]  # 持股数量列表
                    pcts = list_[13: 13 + len(holders)]  # 持股比例列表
                    for i in range(len(holders)):
                        liquid_holder_info = self.liquid_holder_info.filter(code=code).filter(name=holders[i])
                        if liquid_holder_info:
                            liquid_holder_info = liquid_holder_info.get()
                        else:
                            liquid_holder_info = LiquidHolderInfo()
                            liquid_holder_info.code = code
                            liquid_holder_info.name = holders[i]
                            liquid_holder_info.quantity = quantities[i]
                            liquid_holder_info.pct = pcts[i]
                        try:
                            liquid_holder_info.save()
                        except Exception as e:
                            print("LiquidHolderInfo表保存失败，code_name:{}_{}，报错内容：{}".format(code, quantities[i], e))
                except Exception as e:
                    print("数据表更新失败，{}，报错内容：{}".format(row, e))
        return HttpResponseRedirect(reverse('index'))


class MarketDayView(BaseView):
    """日K线"""

    def post(self, request):
        data = self.excel(request, "market_day")
        if data:
            for row in islice(data, 1, None):
                try:  # 此处一定要与excel中字段顺序对应！！
                    code, date, close, open, high, low, volume, pct_change, swing, avg_price, turn, amount, buy_amount, sell_amount, buy_volume, \
                    sell_volume, net_buy_amount, net_buy_volume, active_buy_volume, active_sell_volume, buy_amount_active, sell_amount_active = \
                        map(lambda x: x.value, row)
                    code = code.split(".")[0]
                    market_day = self.market_day.filter(code=code).filter(date=date)
                    if market_day:
                        market_day = market_day.get()
                    else:
                        market_day = MarketDay()
                        market_day.code = code
                        market_day.date = date
                    market_day.close = close
                    market_day.open = open
                    market_day.high = high
                    market_day.low = low
                    market_day.pct_change = pct_change
                    market_day.swing = swing
                    market_day.avg_price = avg_price
                    market_day.turn = turn
                    market_day.volume = volume
                    market_day.amount = amount
                    market_day.buy_amount = buy_amount
                    market_day.sell_amount = sell_amount
                    market_day.buy_volume = buy_volume
                    market_day.sell_volume = sell_volume
                    market_day.net_buy_amount = net_buy_amount
                    market_day.net_buy_volume = net_buy_volume
                    market_day.active_buy_volume = active_buy_volume
                    market_day.active_sell_volume = active_sell_volume
                    market_day.buy_amount_active = buy_amount_active
                    market_day.sell_amount_active = sell_amount_active
                    try:
                        market_day.save()
                    except Exception as e:
                        print("MarketDay表保存失败，code_date:{}_{}，报错内容：{}".format(code, date, e))
                except Exception as e:
                    print("数据表更新失败，{}，报错内容：{}".format(row, e))
        return HttpResponseRedirect(reverse('index'))


class MarketWeekView(BaseView):
    """周K线"""

    def post(self, request):
        return HttpResponseRedirect(reverse('index'))


class MarketMonthView(BaseView):
    """月K线"""

    def post(self, request):
        return HttpResponseRedirect(reverse('index'))


class MarketYearView(BaseView):
    """年K线"""

    def post(self, request):
        return HttpResponseRedirect(reverse('index'))


class MarginView(BaseView):
    """融资融券"""

    def post(self, request):
        data = self.excel(request, "margin")
        if data:
            for row in islice(data, 1, None):
                try:  # 此处一定要与excel中字段顺序对应！！
                    code, date, purchase_with_borrowed_money, repayment_to_broker, trading_balance, sales_of_borrowed_sec, \
                    repayment_of_borrowed_sec, sale_trading_amount, sec_lending_balance_volume, sec_lending_balance, \
                    trading_and_sec_lending_balance = map(lambda x: x.value, row)
                    code = code.split(".")[0]
                    margin = self.margin.filter(code=code).filter(date=date)
                    if margin:
                        margin = margin.get()
                    else:
                        margin = Margin()
                        margin.code = code
                        margin.date = date
                    margin.purchase_with_borrowed_money = purchase_with_borrowed_money
                    margin.repayment_to_broker = repayment_to_broker
                    margin.trading_balance = trading_balance
                    margin.sales_of_borrowed_sec = sales_of_borrowed_sec
                    margin.repayment_of_borrowed_sec = repayment_of_borrowed_sec
                    margin.sale_trading_amount = sale_trading_amount
                    margin.sec_lending_balance_volume = sec_lending_balance_volume
                    margin.sec_lending_balance = sec_lending_balance
                    margin.trading_and_sec_lending_balance = trading_and_sec_lending_balance
                    try:
                        margin.save()
                    except Exception as e:
                        print("Margin表保存失败，code_date:{}_{}，报错内容：{}".format(code, date, e))
                except Exception as e:
                    print("数据表更新失败，{}，报错内容：{}".format(row, e))
        return HttpResponseRedirect(reverse('index'))


class MorphologicFactorView(BaseView):
    """形态因子"""

    def post(self, request):
        return HttpResponseRedirect(reverse('index'))


class TechnicalFactorView(BaseView):
    """技术因子"""

    def post(self, request):
        return HttpResponseRedirect(reverse('index'))


class FinancialFactorView(BaseView):
    """财务因子，每支股票有季报、半年报、年报"""

    def post(self, request):
        data = self.excel(request, "financial_factor")
        if data:
            for row in islice(data, 1, None):
                try:  # 此处一定要与excel中字段顺序对应！！
                    code, date, eps, bps, gross_earnings_per_share, capital_stock_per_share, earnings_per_share, undistributed_profit_per_share, \
                    pre_tax_earnings_per_share, ebitda, pe, roe, roa, roic, gross_profit_margin, net_profit_margin, opt_oebt, debt_to_assets, \
                    long_debt_to_long_captial, short_debt_to_long_captial, assets_to_equity, fa_current, fa_quick, fa_cash_to_current_debt, \
                    fa_debt_to_equity, fa_invturn, fa_arturn, fa_assets_turn = map(lambda x: x.value, row)
                    code = code.split(".")[0]
                    year = date.year
                    time = int(date.month / 3)  # 判断财报时间，1-4分别表示1-4季度财报
                    financial_factor = self.financial_factor.filter(code=code).filter(year=year).filter(time=time)
                    if financial_factor:
                        financial_factor = financial_factor.get()
                    else:
                        financial_factor = FinancialFactor()
                        financial_factor.code = code
                        financial_factor.year = year
                        financial_factor.time = time  # code,year,time是联合唯一键
                    financial_factor.eps = eps
                    financial_factor.bps = bps
                    financial_factor.gross_earnings_per_share = gross_earnings_per_share
                    financial_factor.capital_stock_per_share = capital_stock_per_share
                    financial_factor.earnings_per_share = earnings_per_share
                    financial_factor.undistributed_profit_per_share = undistributed_profit_per_share
                    financial_factor.pre_tax_earnings_per_share = pre_tax_earnings_per_share
                    financial_factor.ebitda = ebitda
                    financial_factor.pe = pe
                    financial_factor.roe = roe
                    financial_factor.roa = roa
                    financial_factor.roic = roic
                    financial_factor.gross_profit_margin = gross_profit_margin
                    financial_factor.net_profit_margin = net_profit_margin
                    financial_factor.opt_oebt = opt_oebt
                    financial_factor.debt_to_assets = debt_to_assets
                    financial_factor.long_debt_to_long_captial = long_debt_to_long_captial
                    financial_factor.short_debt_to_long_captial = short_debt_to_long_captial
                    financial_factor.assets_to_equity = assets_to_equity
                    financial_factor.fa_current = fa_current
                    financial_factor.fa_quick = fa_quick
                    financial_factor.fa_cash_to_current_debt = fa_cash_to_current_debt
                    financial_factor.fa_debt_to_equity = fa_debt_to_equity
                    financial_factor.fa_invturn = fa_invturn
                    financial_factor.fa_arturn = fa_arturn
                    financial_factor.fa_assets_turn = fa_assets_turn
                    try:
                        financial_factor.save()
                    except Exception as e:
                        print("FinancialFactor表保存失败，code_time:{}_{}，报错内容：{}".format(code, date, e))
                except Exception as e:
                    print("数据表更新失败，{}，报错内容：{}".format(row, e))
        return HttpResponseRedirect(reverse('index'))
